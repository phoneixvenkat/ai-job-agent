import os
import pathlib
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.utils.logger import get_logger
from database.mysql_db import (
    get_all_applications, get_stats as mysql_get_stats,
    update_application_status, log_application, save_job,
    get_applications_for_report,
)
log = get_logger('tracker')

ROOT      = pathlib.Path(__file__).parent.parent
EXCEL_DIR = str(ROOT / "artifacts")
os.makedirs(EXCEL_DIR, exist_ok=True)


def get_all_applications_tracker() -> list:
    return get_all_applications()


def log_application_tracker(job: dict, resume_path: str, cover_path: str,
                             status: str = "applied") -> int:
    return log_application(job, resume_path, cover_path, status)


def update_status(app_id: int, status: str):
    update_application_status(app_id, status)


def get_stats() -> dict:
    stats = mysql_get_stats()
    stats.setdefault("skipped", 0)
    return stats


def generate_excel_report() -> str:
    apps  = get_applications_for_report()
    stats = get_stats()
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    path  = os.path.join(EXCEL_DIR, f"job_report_{today}.xlsx")
    wb    = Workbook()

    DARK   = "1F4E79"
    MID    = "2E75B6"
    LIGHT  = "BDD7EE"
    WHITE  = "FFFFFF"
    GRAY   = "F2F2F2"
    GREEN  = "E2EFDA"
    ORANGE = "FCE4D6"
    RED    = "FFDFD7"

    def hfont(color=WHITE, size=10, bold=True):
        return Font(color=color, size=size, bold=bold, name="Calibri")

    def bfont(size=10, bold=False, color="000000"):
        return Font(size=size, bold=bold, name="Calibri", color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "JobPilot AI - Daily Report"
    ws1["A1"].font      = Font(color=WHITE, size=16, bold=True, name="Calibri")
    ws1["A1"].fill      = fill(DARK)
    ws1["A1"].alignment = center()
    ws1.row_dimensions[1].height = 35

    ws1.merge_cells("A2:F2")
    ws1["A2"] = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y %H:%M')}  |  Total: {stats.get('total',0)} applications"
    ws1["A2"].font      = Font(color=WHITE, size=10, name="Calibri")
    ws1["A2"].fill      = fill(MID)
    ws1["A2"].alignment = center()

    stat_data = [
        ["Total Applied", stats.get("applied", 0)],
        ["Skipped",       stats.get("skipped", 0)],
        ["Interviews",    stats.get("interview", 0)],
        ["Offers",        stats.get("offer", 0)],
        ["Avg Fit Score", f"{stats.get('avg_fit', 0)}%"],
        ["Report Date",   today]
    ]
    for i, (label, value) in enumerate(stat_data, start=4):
        ws1[f"A{i}"] = label
        ws1[f"A{i}"].font   = bfont(10, True, DARK)
        ws1[f"A{i}"].fill   = fill(LIGHT)
        ws1[f"A{i}"].border = border()
        ws1[f"B{i}"] = str(value)
        ws1[f"B{i}"].font   = bfont(10)
        ws1[f"B{i}"].border = border()
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 30

    # Sheet 2 — Applications
    ws2 = wb.create_sheet("Applications")
    headers    = ["#","Date","Job Title","Company","Location","Source","Fit %","ATS %","Salary Range","Status","Follow-up","Resume","Cover Letter"]
    col_widths = [4,18,35,20,15,12,8,8,15,12,12,25,25]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws2.cell(row=1, column=col, value=h)
        cell.font      = hfont()
        cell.fill      = fill(DARK)
        cell.alignment = center()
        cell.border    = border()
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 22

    status_fills = {
        "applied":   GREEN,
        "skipped":   ORANGE,
        "interview": "FFF2CC",
        "offer":     "D9F2E6",
        "rejected":  RED
    }

    for i, app in enumerate(apps, start=2):
        salary_range = f"${app.get('salary_min',0)//1000}k-${app.get('salary_max',0)//1000}k"
        row_fill     = fill(GRAY) if i % 2 == 0 else fill(WHITE)
        data = [
            i - 1,
            str(app.get("applied_at", ""))[:16],
            app.get("title", ""),
            app.get("org", ""),
            app.get("location", ""),
            app.get("source", ""),
            app.get("fit_score", 0),
            app.get("ats_score", 0),
            salary_range,
            app.get("status", ""),
            app.get("followup_date", ""),
            os.path.basename(app.get("resume_path", "") or ""),
            os.path.basename(app.get("cover_path", "") or "")
        ]
        for col, value in enumerate(data, 1):
            cell           = ws2.cell(row=i, column=col, value=value)
            cell.font      = bfont()
            cell.alignment = center() if col in [1, 7, 8, 10] else left()
            cell.border    = border()
            if col == 10:
                s = str(value).lower()
                cell.fill = fill(status_fills.get(s, WHITE))
                cell.font = bfont(10, True)
            else:
                cell.fill = row_fill
        ws2.row_dimensions[i].height = 18

    ws2.freeze_panes = "A2"

    # Sheet 3 — Follow-ups
    ws3 = wb.create_sheet("Follow-ups")
    fu_headers = ["Job Title","Company","Applied","Follow-up Date","Done","Status","Notes"]
    fu_widths  = [35,20,15,15,8,12,35]
    for col, (h, w) in enumerate(zip(fu_headers, fu_widths), 1):
        cell           = ws3.cell(row=1, column=col, value=h)
        cell.font      = hfont()
        cell.fill      = fill(MID)
        cell.alignment = center()
        cell.border    = border()
        ws3.column_dimensions[get_column_letter(col)].width = w

    pending = [a for a in apps if a.get("status") == "applied"]
    for i, app in enumerate(pending, start=2):
        row_fill = fill(GRAY) if i % 2 == 0 else fill(WHITE)
        data = [
            app.get("title", ""), app.get("org", ""),
            str(app.get("applied_at", ""))[:10],
            app.get("followup_date", ""),
            "No", app.get("status", ""), ""
        ]
        for col, value in enumerate(data, 1):
            cell           = ws3.cell(row=i, column=col, value=value)
            cell.font      = bfont()
            cell.fill      = row_fill
            cell.border    = border()
            cell.alignment = left()
        ws3.row_dimensions[i].height = 18

    wb.save(path)
    log.info(f"[OK] Excel report saved: {path}")
    return path


if __name__ == "__main__":
    log.info("[OK] Tracker using MySQL")
    path = generate_excel_report()
    log.info(f"[OK] Report generated: {path}")
